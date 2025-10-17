# appScreens.py (ou o arquivo do seu App)
import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import tkinter.font as tkfont
import os
import sys
import csv
import json
import base64
from pathlib import Path
from datetime import datetime
import requests

class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        # --- Diretório base (compatível com PyInstaller) ---
        if getattr(sys, "frozen", False):
            self.base_dir = Path(sys.executable).parent
        else:
            self.base_dir = Path(__file__).resolve().parent

        # Variáveis de layout
        self.initialAppWidth = 1500
        self.initialAppHeight = 630
        self.screenWidth = self.winfo_screenwidth()
        self.screenHeight = self.winfo_screenheight()
        self.appTitle = {"menu":"3N Menu", "verClientes":"3N Clientes"}
        self.appInitialPosX = int((self.screenWidth / 2) - (self.initialAppWidth / 2))
        self.appInitialPosY = int((self.screenHeight / 2) - (self.initialAppHeight / 2))

        # Dados locais
        self.data_dir = self.base_dir / "Data"
        self.data_file = self.data_dir / "clientes.csv"
        self.clients = []  # [{"empresa": str, "vencimento": str(YYYY-MM-DD)}]

        # Janela
        self.title(self.appTitle["menu"])
        ctk.set_appearance_mode("dark")
        self.geometry(f"{self.initialAppWidth}x{self.initialAppHeight}+{self.appInitialPosX}+{self.appInitialPosY}")
        self.minsize(700, 420)
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # Carregar .env local (DadApp.env ou .env na mesma pasta do exe/script)
        try:
            from dotenv import load_dotenv
            for p in (self.base_dir / "DadApp.env", self.base_dir / ".env"):
                if p.exists():
                    load_dotenv(dotenv_path=p, override=False)
        except Exception:
            pass

        self._ensure_data_store()
        self._load_clients()

        # Frames
        self.menu_frame = self.create_menu()
        self.last_frame = None
        self.clients_frame = self.create_clients_view()
        self.show_menu()

    # ---------- UI ----------
    def create_menu(self):
        frame = ctk.CTkFrame(self)
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        container = ctk.CTkFrame(frame)
        container.grid(row=0, column=0, sticky="nsew")
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        card = ctk.CTkFrame(container, corner_radius=12)
        card.grid(row=0, column=0, padx=30, pady=30)
        for i in range(7):
            card.grid_rowconfigure(i, weight=1)
        card.grid_columnconfigure(0, weight=1)

        title = ctk.CTkLabel(card, text="3N", font=ctk.CTkFont(size=36, weight="bold"))
        title.grid(row=0, column=0, padx=30, pady=(30, 6))
        subtitle = ctk.CTkLabel(card, text="Gestão de Licenças", font=ctk.CTkFont(size=16))
        subtitle.grid(row=1, column=0, padx=30, pady=(0, 20))

        ctk.CTkButton(card, width=280, height=46, corner_radius=10,
                      text="Clientes", command=self.show_clients).grid(row=2, column=0, padx=30, pady=8)

        ctk.CTkButton(card, width=200, height=36, corner_radius=8,
                      text="Sincronizar com servidor", command=self.sync_to_server).grid(row=3, column=0, padx=30, pady=(8, 8))

        ctk.CTkButton(card, width=200, height=36, corner_radius=8,
                      text="Enviar p/ GitHub", command=self.push_to_github).grid(row=4, column=0, padx=30, pady=(0, 16))

        ctk.CTkButton(card, width=140, height=34, corner_radius=8,
                      text="Sair", command=self.destroy).grid(row=5, column=0, padx=30, pady=(0, 24))
        return frame

    def show_menu(self):
        if self.last_frame is not None:
            self.last_frame.pack_forget()
        if self.last_frame == self.menu_frame:
            return
        self.title(self.appTitle["menu"])
        self.menu_frame.pack(fill="both", expand=True)
        self.last_frame = self.menu_frame

    def show_clients(self):
        if self.last_frame == self.clients_frame:
            return
        if self.last_frame is not None:
            self.last_frame.pack_forget()
        self.title(self.appTitle["verClientes"])
        self.clients_frame.pack(fill="both", expand=True)
        self.refresh_table()
        self.last_frame = self.clients_frame

    def create_clients_view(self):
        frame = ctk.CTkFrame(self)
        frame.grid_rowconfigure(1, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        # Top bar
        top = ctk.CTkFrame(frame)
        top.grid(row=0, column=0, sticky="ew", padx=10, pady=10)
        top.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(top, text="Clientes (Empresa / Vencimento)", font=ctk.CTkFont(size=18, weight="bold")).grid(row=0, column=0, sticky="w")
        ctk.CTkButton(top, width=120, height=30, corner_radius=6, text="Menu", command=self.show_menu).grid(row=0, column=1, padx=6)
        ctk.CTkButton(top, width=140, height=30, corner_radius=6, text="Sincronizar", command=self.sync_to_server).grid(row=0, column=2, padx=6)
        ctk.CTkButton(top, width=140, height=30, corner_radius=6, text="Enviar p/ GitHub", command=self.push_to_github).grid(row=0, column=3, padx=6)

        # Tabela
        table_frame = ctk.CTkFrame(frame)
        table_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        tv_container = tk.Frame(table_frame)
        tv_container.grid(row=0, column=0, sticky="nsew")
        tv_container.grid_rowconfigure(0, weight=1)
        tv_container.grid_columnconfigure(0, weight=1)

        self.tree = ttk.Treeview(tv_container, columns=("empresa", "vencimento"), show="headings", selectmode="browse")
        self.tree.heading("empresa", text="Empresa")
        self.tree.heading("vencimento", text="Vencimento (DD/MM/AAAA)")
        self.tree.column("empresa", anchor=tk.W, width=520, stretch=True)
        self.tree.column("vencimento", anchor=tk.CENTER, width=200, stretch=True)

        y_scroll = ttk.Scrollbar(tv_container, orient="vertical", command=self.tree.yview)
        x_scroll = ttk.Scrollbar(tv_container, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")

        style = ttk.Style(self)
        try:
            style.theme_use(style.theme_use())
        except Exception:
            pass
        self._tv_font = tkfont.Font(family='Segoe UI', size=13, weight='bold')
        self._tv_head_font = tkfont.Font(family='Segoe UI', size=14, weight='bold')
        style.configure('Treeview', font=self._tv_font, rowheight=30)
        style.configure('Treeview.Heading', font=self._tv_head_font)

        self.tree.tag_configure('expired', background='#7a2f2f', foreground='#ffffff')
        self.tree.tag_configure('due_15', background='#a85f00', foreground='#ffffff')
        self.tree.tag_configure('due_month', background='#b59f3b', foreground='#ffffff')
        self.tree.tag_configure('ok_far', background='#2f7a3d', foreground='#ffffff')
        self.tree.tag_configure('normal', background='')
        self.tree.tag_configure('even', background='#1f1f1f')
        self.tree.tag_configure('odd', background='#262626')

        # Controles
        controls = ctk.CTkFrame(frame)
        controls.grid(row=2, column=0, sticky="ew", padx=10, pady=(0, 10))
        for i in range(10):
            controls.grid_columnconfigure(i, weight=1)

        ctk.CTkLabel(controls, text="Empresa:").grid(row=0, column=0, padx=6, pady=6, sticky="e")
        self.empresa_entry = ctk.CTkEntry(controls, placeholder_text="Nome da empresa")
        self.empresa_entry.grid(row=0, column=1, padx=6, pady=6, sticky="ew")

        ctk.CTkLabel(controls, text="Vencimento:").grid(row=0, column=2, padx=6, pady=6, sticky="e")
        self.venc_entry = ctk.CTkEntry(controls, placeholder_text="DD/MM/AAAA")
        self.venc_entry.grid(row=0, column=3, padx=6, pady=6, sticky="ew")

        ctk.CTkButton(controls, text="Adicionar", width=110, command=self.add_client).grid(row=0, column=4, padx=6, pady=6)
        ctk.CTkButton(controls, text="Remover selecionado", width=160, command=self.remove_selected_client).grid(row=0, column=5, padx=6, pady=6)
        ctk.CTkButton(controls, text="Editar selecionado", width=150, command=self.edit_selected_client).grid(row=0, column=6, padx=6, pady=6)
        ctk.CTkButton(controls, text="Exportar (Excel)", width=120, command=self.export_clients_excel).grid(row=0, column=7, padx=6, pady=6)
        ctk.CTkButton(controls, text="Recarregar", width=110, command=self.refresh_table).grid(row=0, column=8, padx=6, pady=6)
        ctk.CTkButton(controls, text="Enviar p/ GitHub", width=140, command=self.push_to_github).grid(row=0, column=9, padx=6, pady=6)

        ctk.CTkLabel(frame, text="Legenda: Longe (verde), 30 dias (amarelo), 15 dias (laranja), Vencidas (vermelho)").grid(row=3, column=0, padx=12, pady=(0, 10), sticky='w')
        return frame

    # ---------- Dados ----------
    def _ensure_data_store(self):
        self.data_dir.mkdir(parents=True, exist_ok=True)
        if not self.data_file.exists():
            with self.data_file.open("w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["empresa", "vencimento"])  # header

    def _load_clients(self):
        self.clients = []
        try:
            with self.data_file.open("r", newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    emp = (row.get("empresa") or "").strip()
                    ven = (row.get("vencimento") or "").strip()
                    if not emp:
                        continue
                    d = self._parse_date_any(ven)
                    ven_iso = d.strftime("%Y-%m-%d") if d else ven
                    self.clients.append({"empresa": emp, "vencimento": ven_iso})
        except FileNotFoundError:
            self._ensure_data_store()

    def _save_clients(self, commit_message="Update clientes.csv from desktop app"):
        with self.data_file.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["empresa", "vencimento"])
            for c in self.clients:
                writer.writerow([c["empresa"], self._format_date_display(c.get("vencimento", ""))])

        # Envio automático ao GitHub se habilitado
        if (os.environ.get("GITHUB_PUSH_ON_SAVE") or "").strip() == "1":
            try:
                self._push_github_internal(commit_message, silent=True)
            except Exception:
                pass

    # ---------- Ações UI ----------
    def refresh_table(self):
        for i in self.tree.get_children():
            self.tree.delete(i)

        def sort_key(c):
            d = self._parse_date_any(c.get("vencimento", ""))
            if d is not None:
                return (0, d)
            return (1, c.get("empresa", ""))

        for idx, c in enumerate(sorted(self.clients, key=sort_key)):
            tag = self._row_tag_for_client(c)
            display_date = self._format_date_display(c.get("vencimento", ""))
            tags = (('even' if idx % 2 == 0 else 'odd'),) if tag == 'normal' else (tag,)
            self.tree.insert("", tk.END, values=(c.get("empresa",""), display_date), tags=tags)

    def add_client(self):
        empresa = self.empresa_entry.get().strip()
        venc = self.venc_entry.get().strip()
        if not empresa:
            messagebox.showerror("Erro", "Informe o nome da empresa.")
            return
        if venc:
            if not self._valid_date(venc):
                messagebox.showerror("Erro", "Data inválida. Use DD/MM/AAAA.")
                return
            venc_iso = self._to_iso_str(venc)
        else:
            venc_iso = ""
        self.clients.append({"empresa": empresa, "vencimento": venc_iso})
        self._save_clients("Add client from desktop app")
        self.empresa_entry.delete(0, tk.END)
        self.venc_entry.delete(0, tk.END)
        self.refresh_table()

    def remove_selected_client(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Remover", "Selecione um cliente na tabela.")
            return
        values = self.tree.item(sel[0], "values")
        empresa_sel, venc_sel = values[0], values[1]
        venc_sel_iso = self._to_iso_str(venc_sel) if venc_sel else ""
        removed = False
        for i, c in enumerate(self.clients):
            if c.get("empresa") != empresa_sel:
                continue
            stored = c.get("vencimento", "")
            if (venc_sel_iso and stored == venc_sel_iso) or (not venc_sel_iso and self._format_date_display(stored) == venc_sel):
                del self.clients[i]
                removed = True
                break
        if removed:
            self._save_clients("Remove client from desktop app")
            self.refresh_table()
        else:
            messagebox.showwarning("Aviso", "Cliente não encontrado nos dados.")

    def edit_selected_client(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Editar", "Selecione um cliente na tabela.")
            return
        values = self.tree.item(sel[0], "values")
        old_empresa, old_venc_disp = values[0], values[1]
        old_venc_iso = self._to_iso_str(old_venc_disp) if old_venc_disp else ""

        dlg = ctk.CTkToplevel(self)
        dlg.title("Editar Cliente - 3N")
        dlg.geometry("420x200")
        dlg.grab_set()
        dlg.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(dlg, text="Empresa:").grid(row=0, column=0, padx=10, pady=(15, 8), sticky="e")
        empresa_e = ctk.CTkEntry(dlg)
        empresa_e.grid(row=0, column=1, padx=10, pady=(15, 8), sticky="ew")
        empresa_e.insert(0, old_empresa)

        ctk.CTkLabel(dlg, text="Vencimento (DD/MM/AAAA):").grid(row=1, column=0, padx=10, pady=8, sticky="e")
        venc_e = ctk.CTkEntry(dlg, placeholder_text="DD/MM/AAAA")
        venc_e.grid(row=1, column=1, padx=10, pady=8, sticky="ew")
        venc_e.insert(0, old_venc_disp)

        def on_save():
            new_emp = empresa_e.get().strip()
            new_venc = venc_e.get().strip()
            if not new_emp:
                messagebox.showerror("Editar", "Informe o nome da empresa.")
                return
            if new_venc and not self._valid_date(new_venc):
                messagebox.showerror("Editar", "Data inválida. Use DD/MM/AAAA.")
                return
            new_venc_iso = self._to_iso_str(new_venc) if new_venc else ""
            updated = False
            for i, c in enumerate(self.clients):
                if c.get("empresa") != old_empresa:
                    continue
                stored = c.get("vencimento", "")
                if (old_venc_iso and stored == old_venc_iso) or (not old_venc_iso and self._format_date_display(stored) == old_venc_disp):
                    self.clients[i] = {"empresa": new_emp, "vencimento": new_venc_iso}
                    updated = True
                    break
            if not updated:
                messagebox.showwarning("Editar", "Registro original não encontrado. Recarregue a tabela.")
            else:
                self._save_clients("Edit client from desktop app")
                self.refresh_table()
            dlg.destroy()

        btn_row = ctk.CTkFrame(dlg)
        btn_row.grid(row=2, column=0, columnspan=2, pady=15)
        ctk.CTkButton(btn_row, text="Salvar", width=120, command=on_save).pack(side=tk.LEFT, padx=6)
        ctk.CTkButton(btn_row, text="Cancelar", width=120, command=dlg.destroy).pack(side=tk.LEFT, padx=6)

    # ---------- Exportação ----------
    def export_clients_excel(self):
        if not self.clients:
            messagebox.showinfo("Exportar", "Não há dados para exportar.")
            return
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel", "*.xlsx")],
                                                 title="Salvar lista de clientes (Excel)")
        if not file_path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = 'Clientes 3N'

            ws.merge_cells('A1:B1')
            c = ws['A1']
            c.value = "3N - Clientes e Vencimentos"
            c.font = Font(size=18, bold=True)
            c.alignment = Alignment(horizontal='center')

            ws.merge_cells('A2:B2')
            gen = ws['A2']
            gen.value = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            gen.font = Font(size=12)
            gen.alignment = Alignment(horizontal='center')

            ws['A3'].value = 'Empresa'
            ws['B3'].value = 'Vencimento'
            header_fill = PatternFill('solid', fgColor='1F4E78')
            header_font = Font(color='FFFFFF', bold=True, size=14)
            for col in ('A', 'B'):
                cell = ws[f'{col}3']
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            thin = Side(style='thin', color='CCCCCC')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            row_start = 4

            def sort_key(crow):
                d = self._parse_date_any(crow.get('vencimento',''))
                return (0, d) if d else (1, crow.get('empresa',''))

            sorted_rows = sorted(self.clients, key=sort_key)
            for i, cdata in enumerate(sorted_rows):
                empresa = cdata.get('empresa','')
                venc = self._format_date_display(cdata.get('vencimento',''))
                status = self._row_tag_for_client(cdata)
                r = row_start + i
                ws[f'A{r}'].value = empresa
                ws[f'B{r}'].value = venc

                for col in ('A', 'B'):
                    cell = ws[f'{col}{r}']
                    cell.font = Font(size=14)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left' if col=='A' else 'center', vertical='center')

                if status == 'expired':
                    fill = PatternFill('solid', fgColor='FFC7CE')
                elif status == 'due_15':
                    fill = PatternFill('solid', fgColor='FCE4D6')
                elif status == 'due_month':
                    fill = PatternFill('solid', fgColor='FFF2CC')
                elif status == 'ok_far':
                    fill = PatternFill('solid', fgColor='E2EFDA')
                else:
                    fill = None
                if fill:
                    ws[f'A{r}'].fill = fill
                    ws[f'B{r}'].fill = fill
                ws.row_dimensions[r].height = 24

            ws.column_dimensions['A'].width = 60
            ws.column_dimensions['B'].width = 20
            ws.freeze_panes = 'A4'
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToWidth = 1
            ws.sheet_view.showGridLines = False
            ws.print_title_rows = '1:3'
            ws.page_margins.left = ws.page_margins.right = 0.4
            ws.page_margins.top = ws.page_margins.bottom = 0.5

            wb.save(file_path)
            messagebox.showinfo("Exportar", f"Exportado com sucesso para:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Erro ao exportar Excel",
                                 "Falha ao exportar para Excel com formatação. Instale 'openpyxl'.\n\nErro: " + str(e))

    # ---------- Utils ----------
    def _valid_date(self, s):
        try:
            datetime.strptime(s, "%d/%m/%Y")
            return True
        except Exception:
            return False

    def _parse_date_any(self, s):
        if not s:
            return None
        s = s.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                continue
        return None

    def _to_iso_str(self, s):
        d = self._parse_date_any(s)
        if d is None:
            return ""
        return d.strftime("%Y-%m-%d")

    def _format_date_display(self, s):
        d = self._parse_date_any(s)
        if d is None:
            return s or ""
        return d.strftime("%d/%m/%Y")

    def _row_tag_for_client(self, c):
        ven = c.get("vencimento", "").strip()
        if not ven:
            return 'normal'
        d = self._parse_date_any(ven)
        if d is None:
            return 'normal'
        today = datetime.today().date()
        if d < today:
            return 'expired'
        delta = (d - today).days
        if delta <= 15:
            return 'due_15'
        if delta <= 30:
            return 'due_month'
        return 'ok_far'

    # ---------- Sincronização com Render (já existente) ----------
    def sync_to_server(self):
        url = (os.environ.get("RENDER_API_URL") or os.environ.get("RENDER_URL") or os.environ.get("API_URL") or "").strip()
        token = (os.environ.get("RENDER_API_KEY") or os.environ.get("RENDER_API_TOKEN") or os.environ.get("API_KEY") or "").strip()
        if not url:
            messagebox.showwarning("Sincronizar", "Defina RENDER_API_URL no arquivo DadApp.env.")
            return
        payload = {
            "clients": [
                {
                    "empresa": c.get("empresa", ""),
                    "vencimento_iso": (self._parse_date_any(c.get("vencimento", "")) or datetime.today().date()).strftime("%Y-%m-%d") if c.get("vencimento") else "",
                    "vencimento_br": self._format_date_display(c.get("vencimento", "")),
                }
                for c in self.clients
            ],
            "generated_at": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        }
        headers = {"Content-Type": "application/json"}
        if token:
            headers["Authorization"] = f"Bearer {token}"
        try:
            resp = requests.post(url, json=payload, headers=headers, timeout=30)
            if 200 <= resp.status_code < 300:
                messagebox.showinfo("Sincronizar", "Dados enviados com sucesso ao servidor.")
            else:
                messagebox.showerror("Sincronizar", f"Falha ao enviar: {resp.status_code} {resp.text[:200]}")
        except Exception as e:
            messagebox.showerror("Sincronizar", f"Erro de rede: {e}")

    # ---------- Push para GitHub (sem git instalado) ----------
    def push_to_github(self):
        try:
            self._push_github_internal("Sync clientes.csv from desktop app")
            messagebox.showinfo("GitHub", "Arquivo enviado/atualizado com sucesso.")
        except Exception as e:
            messagebox.showerror("GitHub", f"Falha ao enviar: {e}")

    def _push_github_internal(self, commit_message: str, silent=False):
        repo = (os.environ.get("GITHUB_REPO") or "").strip()             # ex: NVdev22/3N-CLIENTES
        token = (os.environ.get("GITHUB_TOKEN") or "").strip()           # Fine-grained token (repo:contents read/write)
        file_path = (os.environ.get("GITHUB_FILE") or "clientes.csv").strip()
        branch = (os.environ.get("GITHUB_BRANCH") or "main").strip()
        committer_name = (os.environ.get("GITHUB_COMMITTER_NAME") or "3N Bot").strip()
        committer_email = (os.environ.get("GITHUB_COMMITTER_EMAIL") or "noreply@local").strip()

        if not repo or not token:
            raise RuntimeError("Defina GITHUB_REPO e GITHUB_TOKEN no DadApp.env.")

        headers = {
            "Accept": "application/vnd.github+json",
            "Authorization": f"Bearer {token}",
            "X-GitHub-Api-Version": "2022-11-28"
        }

        # 1) Descobrir SHA atual (se existir)
        sha = None
        url_get = f"https://api.github.com/repos/{repo}/contents/{file_path}?ref={branch}"
        r = requests.get(url_get, headers=headers, timeout=20)
        if r.status_code == 200:
            sha = r.json().get("sha")
        elif r.status_code not in (404,):
            raise RuntimeError(f"GET {r.status_code}: {r.text[:200]}")

        # 2) Codificar conteúdo local
        content_bytes = self.data_file.read_bytes()
        encoded = base64.b64encode(content_bytes).decode("ascii")

        # 3) PUT (cria/atualiza)
        url_put = f"https://api.github.com/repos/{repo}/contents/{file_path}"
        payload = {
            "message": commit_message,
            "content": encoded,
            "branch": branch,
            "committer": {"name": committer_name, "email": committer_email},
        }
        if sha:
            payload["sha"] = sha

        r2 = requests.put(url_put, headers=headers, data=json.dumps(payload), timeout=25)
        if r2.status_code not in (200, 201):
            raise RuntimeError(f"PUT {r2.status_code}: {r2.text[:200]}")
        if not silent:
            print("GitHub: upload OK")

if __name__ == "__main__":
    app = App()
    app.mainloop()
